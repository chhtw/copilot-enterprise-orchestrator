"""
test_retail_prices.py — 單元測試：retail_prices + xlsx_builder 模組。
"""

from __future__ import annotations

import json
import os

import pytest

os.environ.setdefault("MOCK_MODE", "true")

from orchestrator_app.contracts import CostLineItem, CostStructureOutput, StepStatus
from orchestrator_app.retail_prices import (
    PricedLineItem,
    RetailPrice,
    build_filter,
    pick_best_price,
    resolve_service_name,
    _score_match,
)
from orchestrator_app.xlsx_builder import build_estimate_xlsx


# ======================================================================
# Fixtures
# ======================================================================
def _make_item(**overrides) -> CostLineItem:
    """Helper to create a CostLineItem with defaults."""
    defaults = dict(
        resource_type="azurerm_firewall",
        name="hub-fw",
        display_name="Azure Firewall",
        sku="AZFW_VNet Premium",
        region="eastasia",
        product_id="azure-firewall",
        meter="Standard Deployment",
        unit="1 Hour",
        pricing_tier="Standard",
        quantity=730.0,
        estimated_monthly_usd=912.50,
        notes="",
    )
    defaults.update(overrides)
    return CostLineItem(**defaults)


# ======================================================================
# Test: resolve_service_name
# ======================================================================
class TestResolveServiceName:
    def test_product_slug(self):
        item = _make_item(product_id="azure-firewall")
        assert resolve_service_name(item) == "Azure Firewall"

    def test_terraform_type_fallback(self):
        item = _make_item(product_id="unknown-slug", resource_type="azurerm_firewall")
        assert resolve_service_name(item) == "Azure Firewall"

    def test_arm_type_fallback(self):
        item = _make_item(
            product_id="", resource_type="Microsoft.Network/azureFirewalls"
        )
        assert resolve_service_name(item) == "Azure Firewall"

    def test_free_resource(self):
        item = _make_item(product_id="resource-groups", resource_type="azurerm_resource_group")
        assert resolve_service_name(item) == ""

    def test_unknown_returns_empty(self):
        item = _make_item(product_id="totally-unknown", resource_type="something_custom")
        assert resolve_service_name(item) == ""

    def test_key_vault(self):
        item = _make_item(product_id="key-vault", resource_type="azurerm_key_vault")
        assert resolve_service_name(item) == "Key Vault"

    def test_vpn_gateway(self):
        item = _make_item(product_id="vpn-gateway")
        assert resolve_service_name(item) == "VPN Gateway"


# ======================================================================
# Test: build_filter
# ======================================================================
class TestBuildFilter:
    def test_basic(self):
        f = build_filter("Azure Firewall", "eastasia")
        assert "serviceName eq 'Azure Firewall'" in f
        assert "armRegionName eq 'eastasia'" in f
        assert "priceType eq 'Consumption'" in f

    def test_global_region(self):
        f = build_filter("Azure DNS", "global")
        assert "armRegionName" not in f

    def test_empty_region(self):
        f = build_filter("Azure Firewall", "")
        assert "armRegionName" not in f


# ======================================================================
# Test: _score_match
# ======================================================================
class TestScoreMatch:
    def test_sku_match_scores_high(self):
        item = _make_item(sku="Premium")
        price = {"skuName": "Premium", "meterName": "Standard Deployment", "retailPrice": 1.0, "type": "Consumption", "productName": "Azure Firewall"}
        score = _score_match(item, price)
        assert score > 5  # SKU match (+10) + non-zero price (+2)

    def test_reservation_penalized(self):
        item = _make_item(sku="Premium")
        consumption_price = {"skuName": "Premium", "meterName": "x", "retailPrice": 1.0, "type": "Consumption", "productName": ""}
        reservation_price = {"skuName": "Premium", "meterName": "x", "retailPrice": 1.0, "type": "Reservation", "productName": ""}
        s_c = _score_match(item, consumption_price)
        s_r = _score_match(item, reservation_price)
        assert s_c > s_r  # Consumption should score higher

    def test_zero_price_for_free_resource(self):
        item = _make_item(estimated_monthly_usd=0.0)
        price = {"skuName": "", "meterName": "", "retailPrice": 0.0, "type": "Consumption", "productName": ""}
        score = _score_match(item, price)
        assert score >= 2  # Both zero → +2


# ======================================================================
# Test: pick_best_price
# ======================================================================
class TestPickBestPrice:
    def test_empty_list(self):
        item = _make_item()
        result = pick_best_price(item, [])
        assert result is None

    def test_single_item(self):
        item = _make_item(sku="Premium")
        api_items = [
            {"serviceName": "Azure Firewall", "skuName": "Premium", "meterName": "Standard Deployment",
             "retailPrice": 1.25, "type": "Consumption", "productName": "Azure Firewall",
             "armRegionName": "eastasia", "unitOfMeasure": "1 Hour", "currencyCode": "USD"},
        ]
        result = pick_best_price(item, api_items)
        assert result is not None
        assert isinstance(result, RetailPrice)
        assert result.retail_price == 1.25

    def test_picks_best_match(self):
        item = _make_item(sku="Premium", meter="Standard Deployment")
        api_items = [
            {"serviceName": "Azure Firewall", "skuName": "Standard", "meterName": "Basic",
             "retailPrice": 0.5, "type": "Consumption", "productName": "Azure Firewall",
             "armRegionName": "eastasia", "unitOfMeasure": "1 Hour", "currencyCode": "USD"},
            {"serviceName": "Azure Firewall", "skuName": "Premium", "meterName": "Standard Deployment",
             "retailPrice": 1.25, "type": "Consumption", "productName": "Azure Firewall",
             "armRegionName": "eastasia", "unitOfMeasure": "1 Hour", "currencyCode": "USD"},
        ]
        result = pick_best_price(item, api_items)
        assert result is not None
        assert result.sku_name == "Premium"


# ======================================================================
# Test: PricedLineItem
# ======================================================================
class TestPricedLineItem:
    def test_display_name_from_display(self):
        item = _make_item(display_name="My Firewall", name="fw")
        pi = PricedLineItem(
            line_item=item, retail_price=None,
            unit_price_usd=0, monthly_cost_usd=0, source="free",
        )
        assert pi.display_name == "My Firewall"

    def test_display_name_fallback_to_name(self):
        item = _make_item(display_name="", name="fw")
        pi = PricedLineItem(
            line_item=item, retail_price=None,
            unit_price_usd=0, monthly_cost_usd=0, source="free",
        )
        assert pi.display_name == "fw"


# ======================================================================
# Test: build_estimate_xlsx
# ======================================================================
class TestBuildEstimateXlsx:
    def _make_priced_items(self) -> list[PricedLineItem]:
        """Create sample PricedLineItem list for xlsx testing."""
        items = [
            PricedLineItem(
                line_item=_make_item(display_name="Azure Firewall", quantity=730, estimated_monthly_usd=912.5),
                retail_price=RetailPrice(
                    service_name="Azure Firewall", sku_name="Premium",
                    arm_region_name="eastasia", meter_name="Standard Deployment",
                    product_name="Azure Firewall", unit_of_measure="1 Hour",
                    retail_price=1.25, currency_code="USD", price_type="Consumption",
                ),
                unit_price_usd=1.25,
                monthly_cost_usd=912.5,
                source="retail_api",
            ),
            PricedLineItem(
                line_item=_make_item(
                    display_name="Resource Group", resource_type="azurerm_resource_group",
                    product_id="resource-groups", quantity=1, estimated_monthly_usd=0,
                ),
                retail_price=None,
                unit_price_usd=0.0,
                monthly_cost_usd=0.0,
                source="free",
            ),
        ]
        return items

    def test_returns_bytes(self):
        items = self._make_priced_items()
        result = build_estimate_xlsx(items)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_valid_xlsx(self):
        """Verify the output is a valid xlsx file (starts with PK zip signature)."""
        items = self._make_priced_items()
        result = build_estimate_xlsx(items)
        # xlsx is a zip file → starts with PK (0x504B)
        assert result[:2] == b"PK"

    def test_with_metadata(self):
        items = self._make_priced_items()
        result = build_estimate_xlsx(
            items,
            project_name="test-project",
            region="eastasia",
            currency="USD",
            commitment="PAYG",
        )
        assert isinstance(result, bytes)
        assert len(result) > 100  # should contain data

    def test_empty_items(self):
        result = build_estimate_xlsx([])
        assert isinstance(result, bytes)
        assert result[:2] == b"PK"

    def test_can_be_opened_by_openpyxl(self):
        """Round-trip: generate xlsx → re-open with openpyxl."""
        import io
        from openpyxl import load_workbook

        items = self._make_priced_items()
        xlsx_bytes = build_estimate_xlsx(items, project_name="round-trip-test")
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        assert ws is not None
        # Should have title row, meta row, header row, 2 data rows, total row, source row
        assert ws.max_row >= 7
        wb.close()
