"""
xlsx_builder.py — 用 openpyxl 產生格式化的 estimate.xlsx。

包含：
  - 標題列、日期、來源說明
  - 逐行資源: Resource Name, SKU, Region, Quantity, Unit, Unit Price, Monthly Cost, Source, Notes
  - 合計列
  - 基本格式（粗體標題、數字格式、自動欄寬）
"""

from __future__ import annotations

import logging
from datetime import datetime, timezone
from io import BytesIO
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from .retail_prices import PricedLineItem

logger = logging.getLogger("orchestrator.xlsx_builder")

# Column definitions
COLUMNS = [
    ("Resource Name", 30),
    ("Resource Type", 28),
    ("SKU", 20),
    ("Region", 14),
    ("Quantity", 12),
    ("Unit", 14),
    ("Unit Price (USD)", 17),
    ("Monthly Cost (USD)", 18),
    ("Source", 14),
    ("Notes", 40),
]


def build_estimate_xlsx(
    priced_items: list[PricedLineItem],
    *,
    project_name: str = "",
    region: str = "",
    currency: str = "USD",
    commitment: str = "PAYG",
) -> bytes:
    """
    Generate a formatted estimate.xlsx and return it as bytes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Cost Estimate"

    # ── Styles ──
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    money_fmt = '#,##0.00'
    qty_fmt = '#,##0.00'
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    total_font = Font(bold=True, size=11)
    total_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    # ── Title rows ──
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = f"Azure Cost Estimate — {project_name}" if project_name else "Azure Cost Estimate"
    title_cell.font = Font(bold=True, size=14, color="1F4E79")

    ws.merge_cells("A2:J2")
    meta_cell = ws["A2"]
    meta_cell.value = (
        f"Generated: {now}  |  Region: {region}  |  Currency: {currency}  |  Commitment: {commitment}  |  "
        f"Source: Azure Retail Prices API + LLM estimates"
    )
    meta_cell.font = Font(size=9, italic=True, color="808080")

    ws.append([])  # Row 3 — blank spacer

    # ── Header row (row 4) ──
    header_row = 4
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # ── Data rows ──
    data_start = header_row + 1
    for row_offset, pi in enumerate(priced_items):
        row = data_start + row_offset
        li = pi.line_item

        values = [
            pi.display_name,
            li.resource_type,
            li.sku or "",
            li.region or "",
            li.quantity,
            _unit_label(pi),
            pi.unit_price_usd,
            pi.monthly_cost_usd,
            _source_label(pi.source),
            li.notes or "",
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx == 10))

            # Number formats
            if col_idx == 5:    # Quantity
                cell.number_format = qty_fmt
            elif col_idx in (7, 8):  # Unit Price, Monthly Cost
                cell.number_format = money_fmt

    # ── Total row ──
    total_row = data_start + len(priced_items)
    total_monthly = sum(pi.monthly_cost_usd for pi in priced_items)

    ws.cell(row=total_row, column=1, value="TOTAL").font = total_font
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.fill = total_fill
        cell.border = thin_border
        cell.font = total_font

    total_cell = ws.cell(row=total_row, column=8, value=total_monthly)
    total_cell.number_format = money_fmt
    total_cell.font = total_font
    total_cell.fill = total_fill

    # ── Source attribution row ──
    attr_row = total_row + 2
    ws.merge_cells(f"A{attr_row}:J{attr_row}")
    attr_cell = ws.cell(row=attr_row, column=1)
    api_count = sum(1 for pi in priced_items if pi.source == "retail_api")
    llm_count = sum(1 for pi in priced_items if pi.source == "llm_estimate")
    free_count = sum(1 for pi in priced_items if pi.source == "free")
    attr_cell.value = (
        f"Pricing sources: {api_count} from Azure Retail Prices API, "
        f"{llm_count} from LLM estimate (Step 7B), {free_count} free resources. "
        f"Prices as of {now}. Actual costs may vary."
    )
    attr_cell.font = Font(size=9, italic=True, color="808080")

    # ── Freeze panes ──
    ws.freeze_panes = f"A{data_start}"

    # ── Auto-filter ──
    ws.auto_filter.ref = f"A{header_row}:J{total_row - 1}"

    # ── Serialize ──
    buf = BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()
    logger.info("[XlsxBuilder] Generated estimate.xlsx: %d bytes, %d items, total=$%.2f",
                len(xlsx_bytes), len(priced_items), total_monthly)
    return xlsx_bytes


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _unit_label(pi: PricedLineItem) -> str:
    """Human-readable unit label."""
    if pi.retail_price and pi.retail_price.unit_of_measure:
        return pi.retail_price.unit_of_measure
    return pi.line_item.unit or pi.line_item.meter or ""


def _source_label(source: str) -> str:
    labels = {
        "retail_api": "Azure API",
        "llm_estimate": "LLM Estimate",
        "free": "Free",
    }
    return labels.get(source, source)
